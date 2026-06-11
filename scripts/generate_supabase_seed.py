#!/usr/bin/env python3
"""
Build normalized Supabase seed files from the current map JS data.

Usage (from project root):
    python scripts/generate_supabase_seed.py

Outputs:
    supabase/seed/icbc_sites.csv
    supabase/seed/staff.csv
    supabase/seed/bubele_care_sites.csv
    supabase/seed/preschool_snapshots.csv
    supabase/seed/ploughing_records.csv
    supabase/seed/maize_buyback_records.csv
    supabase/seed/*_template.csv
    supabase/seed/SUPABASE_IMPORT.md
"""

from __future__ import annotations

import csv
import json
import re
import unicodedata
import warnings
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SEED_DIR = ROOT / "supabase" / "seed"
ICBC_JS = ROOT / "data" / "ICBCSites_6.js"
BUBELE_JS = ROOT / "data" / "BubeleCare_10.js"
MIN_PLOUGH_HOURS = 20
PLOUGHING_WORKBOOKS = {
    2023: [
        ROOT / "Ploghing Project Form 2023.xlsx",
        ROOT / "data" / "Ploghing Project Form 2023.xlsx",
        ROOT / "Ploughting Project form 2023.xlsx",
        ROOT / "data" / "Ploughting Project form 2023.xlsx",
    ],
    2024: [
        ROOT / "Ploughing Project Form 2024.xlsx",
        ROOT / "data" / "Ploughing Project Form 2024.xlsx",
        ROOT / "Ploughting Project form 2024.xlsx",
        ROOT / "data" / "Ploughting Project form 2024.xlsx",
    ],
    2025: [
        ROOT / "Ploughing Project Form 2025.xlsx",
        ROOT / "data" / "Ploughing Project Form 2025.xlsx",
        ROOT / "Ploughting Project form 2025.xlsx",
        ROOT / "data" / "Ploughting Project form 2025.xlsx",
    ],
}
MAIZE_WORKBOOKS = [
    ROOT / "data" / "MAIZE BUYBACK PROJECT.xlsx",
    ROOT / "MAIZE BUYBACK PROJECT.xlsx",
    ROOT / "MAIZE SUSTAINABILITY PROJECT 4.xlsx",
    ROOT / "data" / "MAIZE SUSTAINABILITY PROJECT 4.xlsx",
    ROOT / "Maize Sustainability project project 4.xlsx",
    ROOT / "data" / "Maize Sustainability project project 4.xlsx",
]
MAIZE_FARMER_SKIP_FRAGMENTS = (
    "community member",
    "name of icbc",
    "agent's full name",
    "bank",
    "branch code",
    "account number",
    "account holder",
    "total",
    "icbc maize",
    "icbc sites",
)
SITE_SLUG_ALIASES = {
    "esigcaweni": "sigcaweni",
}


def slugify(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def load_js_feature_collection(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1:
        raise ValueError(f"Could not parse GeoJSON from {path}")
    return json.loads(text[start : end + 1])


def normalise_image_path(raw) -> str:
    if raw is None:
        return ""
    return str(raw).replace("\\", "_").replace("/", "_").replace(":", "_").strip()


def parse_number(text) -> float | None:
    if text is None:
        return None
    raw = str(text).strip()
    if not raw or raw.upper() in {"N/A", "NA", "-", "NONE"}:
        return None
    if "NOT PLOUGHED" in raw.upper():
        return None
    match = re.search(r"[\d.]+", raw.replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def parse_preschool_summary(summary: str) -> dict:
    if not summary:
        return {"children_count": None, "teachers_count": None}
    children = re.search(r"Children:\s*(\d+)", summary, re.I)
    teachers = re.search(r"Teachers:\s*(\d+)", summary, re.I)
    return {
        "children_count": int(children.group(1)) if children else None,
        "teachers_count": int(teachers.group(1)) if teachers else None,
    }


def resolve_existing_path(candidates: list[Path]) -> Path | None:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def workbook_site_slug(raw: str) -> str:
    if not raw:
        return ""
    text = str(raw).strip()
    text = re.sub(r"\bICBC\b", "", text, flags=re.I)
    text = re.sub(r"\bIndvuna\b", "", text, flags=re.I)
    text = re.sub(r"\bCommunity Fields\b", "", text, flags=re.I)
    text = text.replace("KaLiba", "Kaliba")
    text = text.replace("KaNcesi", "Kancesi")
    text = re.sub(r"\s+", " ", text).strip(" -")
    slug = slugify(text)
    return SITE_SLUG_ALIASES.get(slug, slug)


def load_workbook_safely(path: Path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        return openpyxl.load_workbook(path, data_only=True)


def as_float(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def as_int(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def parse_ploughing_workbooks(valid_slugs: set[str]) -> list[dict]:
    records: dict[tuple[str, int], dict] = {}

    for year, candidates in PLOUGHING_WORKBOOKS.items():
        workbook_path = resolve_existing_path(candidates)
        if workbook_path is None:
            continue

        wb = load_workbook_safely(workbook_path)
        if "Summaries" not in wb.sheetnames:
            continue

        ws = wb["Summaries"]
        for row in ws.iter_rows(values_only=True):
            raw_name = row[0] if len(row) > 0 else None
            if not isinstance(raw_name, str):
                continue

            name = raw_name.strip()
            if not name or name in {"Ploughing Summaries", "ICBC", "Total"}:
                continue

            slug = workbook_site_slug(name)
            if not slug or slug not in valid_slugs:
                continue

            hours_ploughed = as_float(row[1] if len(row) > 1 else None)
            families_impacted = as_int(row[2] if len(row) > 2 else None)
            hectares_ploughed = as_float(row[3] if len(row) > 3 else None)

            if hours_ploughed == 0 and families_impacted == 0 and hectares_ploughed == 0:
                continue

            key = (slug, year)
            existing = records.get(key)
            if existing is None:
                records[key] = {
                    "site_slug": slug,
                    "year": year,
                    "hours_ploughed": hours_ploughed,
                    "families_impacted": families_impacted,
                    "hectares_ploughed": hectares_ploughed,
                    "notes": "",
                    "_source_names": [name],
                }
                continue

            existing["hours_ploughed"] = as_float(existing["hours_ploughed"]) + hours_ploughed
            existing["families_impacted"] = as_int(existing["families_impacted"]) + families_impacted
            existing["hectares_ploughed"] = as_float(existing["hectares_ploughed"]) + hectares_ploughed
            existing["_source_names"].append(name)

    final_rows: list[dict] = []
    for row in records.values():
        source_names = row.pop("_source_names", [])
        if len(source_names) > 1:
            row["notes"] = f"Aggregated workbook rows: {', '.join(source_names)}"
        final_rows.append(row)

    final_rows.sort(key=lambda item: (item["year"], item["site_slug"]))
    return final_rows


def maize_workbook_year(wb) -> int:
    for ws in wb.worksheets:
        if ws.title.strip().upper() != "SUMMARIES":
            continue
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            first = row[0] if row else None
            if not isinstance(first, str):
                continue
            match = re.search(r"(20\d{2})", first)
            if match:
                return int(match.group(1))
    return 2025


def find_maize_summaries_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().upper() == "SUMMARIES":
            return wb[name]
    return None


def parse_maize_summaries_sheet(ws, valid_slugs: set[str]) -> dict[str, dict]:
    """Authoritative kg totals per site from the SUMMARIES tab."""
    totals: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        raw_name = row[0] if len(row) > 0 else None
        if not isinstance(raw_name, str):
            continue
        name = raw_name.strip()
        if not name or name.lower() in {"icbc sites", "total"}:
            continue
        if "maize" in name.lower() and "project" in name.lower():
            continue

        slug = workbook_site_slug(name)
        if not slug or slug not in valid_slugs:
            continue

        kilograms_bought = as_float(row[1] if len(row) > 1 else None)
        if kilograms_bought <= 0:
            continue

        totals[slug] = {
            "site_slug": slug,
            "kilograms_bought": kilograms_bought,
            "source_name": name,
        }
    return totals


def find_maize_kg_column(header_row: tuple) -> int | None:
    for index, cell in enumerate(header_row):
        if isinstance(cell, str) and "kilogram" in cell.lower():
            return index
    return None


def is_maize_farmer_name(name: str) -> bool:
    cleaned = re.sub(r"\s+", " ", str(name or "").strip())
    if not cleaned or len(cleaned) > 60:
        return False
    lower = cleaned.lower()
    if any(fragment in lower for fragment in MAIZE_FARMER_SKIP_FRAGMENTS):
        return False
    if cleaned.replace(" ", "").isdigit():
        return False
    if re.fullmatch(r"[\d./\-]+", cleaned):
        return False
    return True


def count_maize_farmers_from_sheet(ws) -> int:
    farmers: set[str] = set()
    kg_column: int | None = None
    seen_header = False

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        first_cell = cells[0] if cells else None
        if not seen_header:
            if isinstance(first_cell, str) and "Community Member" in first_cell:
                seen_header = True
                kg_column = find_maize_kg_column(tuple(cells))
            continue

        farmer_name = str(first_cell or "").strip()
        if not is_maize_farmer_name(farmer_name):
            continue

        kg = 0.0
        if kg_column is not None and kg_column < len(cells):
            kg = as_float(cells[kg_column])
        elif len(cells) > 5:
            kg = as_float(cells[5])

        # Skip metadata rows where a bank account number landed in the kg column.
        if kg <= 0 or kg > 2500:
            continue

        farmers.add(re.sub(r"\s+", " ", farmer_name).lower())

    return len(farmers)


def parse_maize_workbook_rows(valid_slugs: set[str]) -> list[dict]:
    workbook_path = resolve_existing_path(MAIZE_WORKBOOKS)
    if workbook_path is None:
        return []

    wb = load_workbook_safely(workbook_path)
    year = maize_workbook_year(wb)
    summary_ws = find_maize_summaries_sheet(wb)
    summary_totals = (
        parse_maize_summaries_sheet(summary_ws, valid_slugs) if summary_ws is not None else {}
    )

    rows_by_slug: dict[str, dict] = {}

    for ws in wb.worksheets:
        if ws.title.strip().upper() == "SUMMARIES":
            continue

        raw_site_name = ""
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            label = row[0] if len(row) > 0 else None
            if isinstance(label, str) and label.strip().lower() == "name of icbc:":
                raw_site_name = str(row[1] or "").strip()
                break

        if not raw_site_name:
            raw_site_name = ws.title

        slug = workbook_site_slug(raw_site_name)
        if not slug or slug not in valid_slugs:
            continue

        farmers_impacted = count_maize_farmers_from_sheet(ws)
        summary = summary_totals.get(slug)
        kilograms_bought = as_float(summary["kilograms_bought"]) if summary else 0.0

        if farmers_impacted == 0 and kilograms_bought == 0:
            continue

        rows_by_slug[slug] = {
            "site_slug": slug,
            "year": year,
            "farmers_impacted": farmers_impacted,
            "kilograms_bought": round(kilograms_bought, 2),
            "meals_made": 0,
            "notes": (
                f"From {workbook_path.name} / {summary['source_name']}"
                if summary
                else f"From {workbook_path.name}"
            ),
        }

    # Include summary-only sites if a detail tab was missing.
    for slug, summary in summary_totals.items():
        if slug in rows_by_slug:
            continue
        rows_by_slug[slug] = {
            "site_slug": slug,
            "year": year,
            "farmers_impacted": 0,
            "kilograms_bought": round(as_float(summary["kilograms_bought"]), 2),
            "meals_made": 0,
            "notes": f"From {workbook_path.name} / {summary['source_name']} (summary only)",
        }

    rows = list(rows_by_slug.values())
    rows.sort(key=lambda item: (item["year"], item["site_slug"]))
    return rows


def merge_rows(base_rows: list[dict], workbook_rows: list[dict]) -> list[dict]:
    merged = {(row["site_slug"], int(row["year"])): row for row in base_rows}
    for row in workbook_rows:
        merged[(row["site_slug"], int(row["year"]))] = row
    return sorted(merged.values(), key=lambda item: (int(item["year"]), item["site_slug"]))


def ploughing_year(props: dict) -> int:
    raw = (props.get("Ploughing Season") or "").strip()
    if raw.isdigit():
        return int(raw)
    return 2024


def csv_escape_row(row: dict, fieldnames: list[str]) -> dict:
    escaped = {}
    for key in fieldnames:
        value = row.get(key, "")
        if value is None:
            value = ""
        escaped[key] = str(value).replace("\r\n", "\n")
    return escaped


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(csv_escape_row(row, fieldnames))


def write_template(path: Path, fieldnames: list[str], example_row: dict | None = None) -> None:
    rows = [example_row] if example_row else []
    write_csv(path, fieldnames, rows)


def build_import_guide() -> str:
    return """# Import ICBC data into Supabase

This project uses a normalized schema:
- `icbc_sites` stores core site profile data only
- staff and programme history live in separate tables
- CSV imports that use `site_slug` go into staging tables first

## 1. Create tables

Run `schema.sql` in Supabase SQL Editor.

If you already created an older database with ploughing / maize fields on `icbc_sites`,
run `programmes_setup.sql` after `schema.sql` to clean that up.

## 2. Import direct CSVs

Use Supabase Table Editor → Import CSV for these tables:

| Table | CSV file |
|-------|----------|
| `icbc_sites` | `seed/icbc_sites.csv` |
| `bubele_care_sites` | `seed/bubele_care_sites.csv` |
| `staff_import` | `seed/staff.csv` |
| `preschool_snapshots_import` | `seed/preschool_snapshots.csv` |
| `ploughing_records_import` | `seed/ploughing_records.csv` |
| `maize_buyback_records_import` | `seed/maize_buyback_records.csv` |

For staff photos, leave `photo_url` blank until your Streamlit uploader sends the image
to GitHub and writes the final GitHub URL back to Supabase.

## 3. Process staging imports

Run this once after the CSV imports:

```sql
select public.process_all_site_imports();
```

That will move rows from the staging import tables into:
- `staff`
- `preschool_snapshots`
- `ploughing_records`
- `maize_buyback_records`

## 4. Weekly stats

Run `weekly_stats_setup.sql`, then import your spreadsheet CSV into `weekly_stats_import`,
then run:

```sql
select * from public.process_weekly_stats_import();
```

## 5. Regenerate seed files after editing map JS

```bash
python scripts/generate_supabase_seed.py
```
"""


def main() -> None:
    SEED_DIR.mkdir(parents=True, exist_ok=True)

    icbc_data = load_js_feature_collection(ICBC_JS)
    bubele_data = load_js_feature_collection(BUBELE_JS)

    icbc_rows: list[dict] = []
    staff_rows: list[dict] = []
    preschool_rows: list[dict] = []
    ploughing_rows: list[dict] = []
    maize_rows: list[dict] = []

    for feature in icbc_data.get("features", []):
        props = feature.get("properties") or {}
        geom = feature.get("geometry") or {}
        coords = geom.get("coordinates") or [None, None]

        name = (props.get("Site Name") or "").strip()
        if not name:
            continue

        slug = slugify(name)
        lon, lat = None, None
        if isinstance(coords, (list, tuple)) and len(coords) >= 2:
            lon, lat = coords[0], coords[1]

        icbc_rows.append(
            {
                "slug": slug,
                "name": name,
                "region": props.get("Region") or "",
                "year_constructed": (props.get("Year Constructed") or "").strip(),
                "water_source": props.get("Water Source") or "",
                "projects": props.get("Projects") or "",
                "about": props.get("About") or "",
                "site_link": (props.get("Site Link") or "").strip(),
                "video_url": "",
                "photo_path": normalise_image_path(props.get("Photo")),
                "cover_image_url": "",
                "latitude": lat if lat is not None else "",
                "longitude": lon if lon is not None else "",
            }
        )

        year = ploughing_year(props)
        preschool_summary = (props.get("PreSchool") or "").strip()
        parsed_preschool = parse_preschool_summary(preschool_summary)
        if parsed_preschool.get("children_count"):
            preschool_rows.append(
                {
                    "site_slug": slug,
                    "children_count": parsed_preschool["children_count"],
                    "snapshot_date": f"{year}-06-01",
                    "year": year,
                    "teachers_count": parsed_preschool.get("teachers_count") or "",
                    "children_impacted_since_inception": "",
                    "notes": preschool_summary,
                }
            )

        hours = parse_number(props.get("Hours Ploughed"))
        if hours is not None and hours >= MIN_PLOUGH_HOURS:
            families = parse_number(props.get("Families Ploughed for")) or 0
            hectares = parse_number(props.get("Area Ploughed")) or 0
            ploughing_rows.append(
                {
                    "site_slug": slug,
                    "year": year,
                    "hours_ploughed": hours,
                    "families_impacted": int(families),
                    "hectares_ploughed": hectares,
                    "notes": "",
                }
            )

        pastor_name = (props.get("Pastor") or "").strip()
        if pastor_name:
            staff_rows.append(
                {
                    "site_slug": slug,
                    "full_name": pastor_name,
                    "description": "",
                    "role": "Pastor",
                    "year_joined": "",
                    "photo_url": "",
                    "sort_order": "0",
                }
            )

    bubele_rows: list[dict] = []
    for feature in bubele_data.get("features", []):
        props = feature.get("properties") or {}
        geom = feature.get("geometry") or {}
        coords = geom.get("coordinates") or [None, None]

        site_name = (props.get("Site") or "").strip()
        family_code = (props.get("Family Code") or "").strip()
        if not site_name or not family_code:
            continue

        lon, lat = None, None
        if isinstance(coords, (list, tuple)) and len(coords) >= 2:
            lon, lat = coords[0], coords[1]

        bubele_rows.append(
            {
                "site_name": site_name,
                "family_code": family_code,
                "region": props.get("Region") or "",
                "latitude": lat if lat is not None else "",
                "longitude": lon if lon is not None else "",
            }
        )

    icbc_fields = [
        "slug",
        "name",
        "region",
        "year_constructed",
        "water_source",
        "projects",
        "about",
        "site_link",
        "video_url",
        "photo_path",
        "cover_image_url",
        "latitude",
        "longitude",
    ]
    staff_fields = [
        "site_slug",
        "full_name",
        "description",
        "role",
        "year_joined",
        "photo_url",
        "sort_order",
    ]
    preschool_fields = [
        "site_slug",
        "children_count",
        "snapshot_date",
        "year",
        "teachers_count",
        "children_impacted_since_inception",
        "notes",
    ]
    ploughing_fields = [
        "site_slug",
        "year",
        "hours_ploughed",
        "families_impacted",
        "hectares_ploughed",
        "notes",
    ]
    maize_fields = [
        "site_slug",
        "year",
        "farmers_impacted",
        "kilograms_bought",
        "meals_made",
        "notes",
    ]
    bubele_fields = ["site_name", "family_code", "region", "latitude", "longitude"]
    weekly_stats_fields = [
        "church",
        "date",
        "total_attendance",
        "men",
        "women",
        "youth",
        "children",
        "salvations",
        "baptisms",
        "home_visits",
        "meals_food_packs",
        "preschool_attendance",
    ]

    valid_site_slugs = {row["slug"] for row in icbc_rows}
    workbook_ploughing_rows = parse_ploughing_workbooks(valid_site_slugs)
    workbook_maize_rows = parse_maize_workbook_rows(valid_site_slugs)

    final_ploughing_rows = merge_rows(ploughing_rows, workbook_ploughing_rows)
    final_maize_rows = merge_rows(maize_rows, workbook_maize_rows)

    write_csv(SEED_DIR / "icbc_sites.csv", icbc_fields, icbc_rows)
    write_csv(SEED_DIR / "staff.csv", staff_fields, staff_rows)
    write_csv(SEED_DIR / "bubele_care_sites.csv", bubele_fields, bubele_rows)
    write_csv(SEED_DIR / "preschool_snapshots.csv", preschool_fields, preschool_rows)
    write_csv(SEED_DIR / "ploughing_records.csv", ploughing_fields, final_ploughing_rows)
    write_csv(SEED_DIR / "maize_buyback_records.csv", maize_fields, final_maize_rows)

    write_template(
        SEED_DIR / "staff_template.csv",
        staff_fields,
        {
            "site_slug": "bulunga",
            "full_name": "Ndumiso Gwebu",
            "description": "Lead pastor serving the Bulunga community.",
            "role": "Pastor",
            "year_joined": "2021",
            "photo_url": "",
            "sort_order": "0",
        },
    )
    write_template(
        SEED_DIR / "preschool_snapshots_template.csv",
        preschool_fields,
        {
            "site_slug": "bulunga",
            "children_count": "27",
            "snapshot_date": "2024-06-01",
            "year": "2024",
            "teachers_count": "2",
            "children_impacted_since_inception": "",
            "notes": "Teachers: 2, Children: 27",
        },
    )
    write_template(
        SEED_DIR / "ploughing_records_template.csv",
        ploughing_fields,
        {
            "site_slug": "bulunga",
            "year": "2024",
            "hours_ploughed": "50",
            "families_impacted": "31",
            "hectares_ploughed": "75",
            "notes": "",
        },
    )
    write_template(
        SEED_DIR / "maize_buyback_records_template.csv",
        maize_fields,
        {
            "site_slug": "bulunga",
            "year": "2024",
            "farmers_impacted": "12",
            "kilograms_bought": "5000",
            "meals_made": "15000",
            "notes": "",
        },
    )
    write_template(
        SEED_DIR / "weekly_stats_template.csv",
        weekly_stats_fields,
        {
            "church": "Bulunga",
            "date": "2024-12-29",
            "total_attendance": "0",
            "men": "0",
            "women": "0",
            "youth": "0",
            "children": "0",
            "salvations": "0",
            "baptisms": "0",
            "home_visits": "0",
            "meals_food_packs": "0",
            "preschool_attendance": "0",
        },
    )

    (SEED_DIR / "import_order.txt").write_text(
        "See SUPABASE_IMPORT.md for the full normalized import checklist.\n",
        encoding="utf-8",
    )
    (SEED_DIR / "SUPABASE_IMPORT.md").write_text(build_import_guide(), encoding="utf-8")

    print(f"Wrote {len(icbc_rows)} ICBC sites -> {SEED_DIR / 'icbc_sites.csv'}")
    print(f"Wrote {len(staff_rows)} staff rows -> {SEED_DIR / 'staff.csv'}")
    print(f"Wrote {len(preschool_rows)} preschool rows -> {SEED_DIR / 'preschool_snapshots.csv'}")
    print(f"Wrote {len(final_ploughing_rows)} ploughing rows -> {SEED_DIR / 'ploughing_records.csv'}")
    print(f"Wrote {len(final_maize_rows)} maize rows -> {SEED_DIR / 'maize_buyback_records.csv'}")
    print(f"Wrote {len(bubele_rows)} Bubele rows -> {SEED_DIR / 'bubele_care_sites.csv'}")
    print(f"Guide -> {SEED_DIR / 'SUPABASE_IMPORT.md'}")


if __name__ == "__main__":
    main()
