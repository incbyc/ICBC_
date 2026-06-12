#!/usr/bin/env python3
"""Build teacher counts from ICBC Teachers.xlsx (counting only — not staff bios).

Usage (from project root):
    python scripts/excel_to_icbc_teachers.py

Outputs:
    data/icbc_teachers.js — static map teacher roll-up (GitHub Pages + preschool counts)
    supabase/seed/preschool_enrollment.csv — teachers_count per site (latest year)
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import tempfile
import unicodedata
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SEED_DIR = ROOT / "supabase" / "seed"
DATA_JS = ROOT / "data" / "icbc_teachers.js"
STAFF_CSV = SEED_DIR / "staff.csv"
PRESCHOOL_CSV = SEED_DIR / "preschool_enrollment.csv"
TEACHER_WORKBOOKS = [
    ROOT / "ICBC Teachers.xlsx",
    ROOT / "ICBC teacher.xlsx",
    ROOT / "data" / "ICBC Teachers.xlsx",
    ROOT / "data" / "ICBC teacher.xlsx",
]
SITE_SLUG_ALIASES = {
    "esigcaweni": "sigcaweni",
    "ka-liba": "kaliba",
    "ka-ncesi": "kancesi",
}


def slugify(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    slug = text.strip("-")
    return SITE_SLUG_ALIASES.get(slug, slug)


def workbook_site_slug(raw: str) -> str:
    if not raw:
        return ""
    text = str(raw).strip()
    text = re.sub(r"\bICBC\b", "", text, flags=re.I)
    text = text.replace("KaLiba", "Kaliba").replace("KaNcesi", "Kancesi")
    text = text.replace("Ka-Liba", "Kaliba").replace("Ka-Ncesi", "Kancesi")
    text = re.sub(r"\s+", " ", text).strip(" -")
    slug = slugify(text)
    return SITE_SLUG_ALIASES.get(slug, slug)


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name or "").strip().lower())


def format_contact(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def resolve_workbook() -> Path | None:
    for candidate in TEACHER_WORKBOOKS:
        if candidate.is_file():
            return candidate
    return None


def load_valid_site_slugs() -> set[str]:
    path = SEED_DIR / "icbc_sites.csv"
    if not path.is_file():
        return set()
    with path.open(encoding="utf-8", newline="") as handle:
        return {row["slug"] for row in csv.DictReader(handle) if row.get("slug")}


def open_teacher_workbook(path: Path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        try:
            return openpyxl.load_workbook(path, read_only=True, data_only=True)
        except PermissionError:
            temp_copy = Path(tempfile.gettempdir()) / "icbc_teachers_import.xlsx"
            shutil.copy2(path, temp_copy)
            return openpyxl.load_workbook(temp_copy, read_only=True, data_only=True)


def parse_teacher_workbook(path: Path, valid_slugs: set[str]) -> dict[str, list[dict[str, str]]]:
    wb = open_teacher_workbook(path)

    ws = wb[wb.sheetnames[0]]
    by_site: dict[str, list[dict[str, str]]] = defaultdict(list)
    issues: list[str] = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip().strip(",")
        if not name or name.lower() in {"teachers name", "teachers contact details"}:
            continue
        contact = format_contact(row[1] if len(row) > 1 else "")
        town = row[2] if len(row) > 2 else ""
        slug = workbook_site_slug(town)
        if not slug:
            issues.append(f"Missing town for teacher {name!r}")
            continue
        if valid_slugs and slug not in valid_slugs:
            issues.append(f"Unknown site town {town!r} -> {slug} for {name!r}")
            continue
        by_site[slug].append(
            {
                "name": name,
                "contact": contact,
                "town": str(town or "").strip(),
            }
        )

    wb.close()
    if issues:
        print("Warnings:")
        for issue in issues[:20]:
            print(f"  - {issue}")
        if len(issues) > 20:
            print(f"  ... and {len(issues) - 20} more")
    return by_site


def write_teachers_js(by_site: dict[str, list[dict[str, str]]]) -> None:
    payload = {
        slug: {
            "count": len(teachers),
            "teachers": [
                {
                    "full_name": teacher["name"],
                    "role": "Teacher",
                    "description": f"Pre-school teacher at {teacher['town'] or slug} ICBC.",
                }
                for teacher in teachers
            ],
        }
        for slug, teachers in sorted(by_site.items())
    }
    total = sum(entry["count"] for entry in payload.values())
    DATA_JS.parent.mkdir(parents=True, exist_ok=True)
    DATA_JS.write_text(
        "var ICBC_TEACHERS_BY_SITE = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\nvar ICBC_TEACHERS_TOTAL = "
        + str(total)
        + ";\n",
        encoding="utf-8",
    )


def remove_spreadsheet_teacher_rows_from_staff_csv() -> int:
    """Drop short workbook-only Teacher rows (not questionnaire bios)."""
    if not STAFF_CSV.is_file():
        return 0
    with STAFF_CSV.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
        fieldnames = list(rows[0].keys()) if rows else []
    kept = []
    removed = 0
    pattern = re.compile(r"^Pre-school teacher at .+ ICBC\.$")
    for row in rows:
        desc = (row.get("description") or "").strip()
        if row.get("role") == "Teacher" and pattern.match(desc):
            removed += 1
            continue
        kept.append(row)
    if removed:
        with STAFF_CSV.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(kept)
    return removed


def update_preschool_teacher_counts(by_site: dict[str, list[dict[str, str]]]) -> int:
    if not PRESCHOOL_CSV.is_file():
        return 0
    with PRESCHOOL_CSV.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
        fieldnames = list(rows[0].keys()) if rows else [
            "site_slug",
            "year",
            "children_count",
            "teachers_count",
            "notes",
        ]

    latest_year_by_site: dict[str, int] = {}
    for row in rows:
        slug = row.get("site_slug", "")
        year = int(float(row["year"])) if row.get("year") else 0
        if slug:
            latest_year_by_site[slug] = max(latest_year_by_site.get(slug, 0), year)

    updated = 0
    for row in rows:
        slug = row.get("site_slug", "")
        year = int(float(row["year"])) if row.get("year") else 0
        if not slug or year != latest_year_by_site.get(slug):
            continue
        count = len(by_site.get(slug, []))
        if count <= 0:
            continue
        row["teachers_count"] = str(count)
        note = (row.get("notes") or "").strip()
        suffix = "Teachers: ICBC Teachers.xlsx"
        if suffix not in note:
            row["notes"] = (note + (" · " if note else "") + suffix).strip()
        updated += 1

    with PRESCHOOL_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return updated


def main() -> None:
    workbook = resolve_workbook()
    if not workbook:
        raise SystemExit(
            "ICBC Teachers.xlsx not found. Place it in the project root or data/ folder."
        )

    valid_slugs = load_valid_site_slugs()
    by_site = parse_teacher_workbook(workbook, valid_slugs)
    total = sum(len(teachers) for teachers in by_site.values())
    write_teachers_js(by_site)
    staff_removed = remove_spreadsheet_teacher_rows_from_staff_csv()
    preschool_updated = update_preschool_teacher_counts(by_site)

    print(f"Workbook: {workbook}")
    print(f"Sites with teachers: {len(by_site)}")
    print(f"Total teachers: {total}")
    print(f"Wrote {DATA_JS}")
    if staff_removed:
        print(f"Removed {staff_removed} workbook-only teacher row(s) from {STAFF_CSV}")
    print(f"Updated teachers_count on {preschool_updated} preschool_enrollment row(s)")


if __name__ == "__main__":
    main()
