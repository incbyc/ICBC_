#!/usr/bin/env python3
"""
Convert ICBC weekly-stats workbooks into a CSV for Supabase import.

Supports:
  - Consolidated workbook (``ICBC_Church_Stats_Consolidated*.xlsx``) with a
    tabular ``Weekly Data`` sheet (preferred).
  - Legacy per-church block layout (parsed via ``backend/icbc/stats_parser.py``).

Output columns match ``weekly_stats_import``:
  church, date, total_attendance, men, women, youth, children,
  salvations, baptisms, home_visits, meals_food_packs, preschool_attendance

Usage:
    python scripts/excel_to_weekly_stats_csv.py path/to/workbook.xlsx
    python scripts/excel_to_weekly_stats_csv.py path/to/workbook.xlsx -o supabase/seed/weekly_stats.csv
"""

from __future__ import annotations

import argparse
import csv
import importlib.util
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PARSER_PATH = ROOT / "backend" / "icbc" / "stats_parser.py"

CONSOLIDATED_SHEET = "Weekly Data"

FIELDS = [
    "church",
    "date",
    "total_attendance",
    "men",
    "women",
    "youth",
    "children",
    "salvations",
    "baptisms",
    "home_visits",
    "meals_food_packs",
    "preschool_attendance",
]

# Header labels on the consolidated ``Weekly Data`` sheet -> CSV field
CONSOLIDATED_HEADER_ALIASES = {
    "church": "church",
    "date": "date",
    "total attendance": "total_attendance",
    "men": "men",
    "women": "women",
    "youth": "youth",
    "children": "children",
    "salvations": "salvations",
    "baptisms": "baptisms",
    "home visits": "home_visits",
    "meals / food packs": "meals_food_packs",
    "preschool attendance": "preschool_attendance",
}


def load_parser():
    spec = importlib.util.spec_from_file_location("stats_parser", PARSER_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {PARSER_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _format_date(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if " " in text:
        text = text.split(" ", 1)[0]
    return text


def _normalise_header(cell) -> str:
    if cell is None:
        return ""
    return str(cell).strip().lower()


def _map_consolidated_headers(header_row: Iterable) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        key = CONSOLIDATED_HEADER_ALIASES.get(_normalise_header(cell))
        if key:
            mapping[idx] = key
    required = {"church", "date", "total_attendance"}
    if not required.issubset(set(mapping.values())):
        return {}
    return mapping


def parse_consolidated_workbook(workbook: Path) -> Optional[List[dict]]:
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    try:
        if CONSOLIDATED_SHEET not in wb.sheetnames:
            return None

        ws = wb[CONSOLIDATED_SHEET]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []

        col_map = _map_consolidated_headers(header_row)
        if not col_map:
            return None

        seen: Dict[str, dict] = {}
        duplicates = 0

        for row in rows:
            rec = {field: 0 for field in FIELDS}
            for col_idx, field in col_map.items():
                value = row[col_idx] if col_idx < len(row) else None
                if field in ("church", "date"):
                    if field == "church":
                        rec[field] = str(value).strip() if value else ""
                    else:
                        rec[field] = _format_date(value)
                else:
                    rec[field] = _to_int(value)

            if not rec["church"] or not rec["date"]:
                continue

            dedup_key = f"{rec['church'].lower()}|{rec['date']}"
            if dedup_key in seen:
                duplicates += 1
            seen[dedup_key] = rec

        records = list(seen.values())
        records.sort(key=lambda r: (r["church"].lower(), r["date"]))
        return records
    finally:
        wb.close()


def parse_legacy_workbook(workbook: Path) -> Tuple[List[dict], int]:
    stats_parser = load_parser()
    file_bytes = workbook.read_bytes()
    raw_records, duplicates = stats_parser.parse_workbook(file_bytes, workbook.name)

    records: List[dict] = []
    for rec in raw_records:
        records.append(
            {
                "church": rec.get("Church", ""),
                "date": rec.get("Date", ""),
                "total_attendance": _to_int(rec.get("Total Attendance")),
                "men": _to_int(rec.get("Men")),
                "women": _to_int(rec.get("Women")),
                "youth": _to_int(rec.get("Youth")),
                "children": _to_int(rec.get("Children")),
                "salvations": _to_int(rec.get("Salvations")),
                "baptisms": _to_int(rec.get("Baptisms")),
                "home_visits": _to_int(rec.get("Home Visits")),
                "meals_food_packs": _to_int(rec.get("Meals / Food Packs")),
                "preschool_attendance": _to_int(rec.get("Preschool Attendance")),
            }
        )
    records.sort(key=lambda r: (r["church"].lower(), r["date"]))
    return records, duplicates


def write_csv(records: List[dict], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)


def main() -> int:
    ap = argparse.ArgumentParser(description="Excel workbook -> weekly_stats CSV")
    ap.add_argument("workbook", type=Path, help="Path to .xlsx file")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=ROOT / "supabase" / "seed" / "weekly_stats.csv",
        help="Output CSV path",
    )
    args = ap.parse_args()

    if not args.workbook.exists():
        print(f"File not found: {args.workbook}", file=sys.stderr)
        return 1

    duplicates = 0
    source = "legacy block layout"

    consolidated = parse_consolidated_workbook(args.workbook)
    if consolidated is not None:
        records = consolidated
        source = f'"{CONSOLIDATED_SHEET}" sheet'
    else:
        records, duplicates = parse_legacy_workbook(args.workbook)

    write_csv(records, args.output)

    print(f"Parsed from {source}")
    print(f"Wrote {len(records)} rows -> {args.output}")
    if duplicates:
        print(f"Duplicates removed inside workbook: {duplicates}")
    print()
    print("Next in Supabase:")
    print("  1. Run weekly_stats_setup.sql (once)")
    print("  2. Table Editor -> weekly_stats_import -> Import CSV")
    print("  3. SQL Editor -> select * from process_weekly_stats_import();")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
