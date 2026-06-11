#!/usr/bin/env python3
"""Build preschool_enrollment.csv from Enrollment_Counts.xlsx.

Usage (from project root):
    python scripts/import_preschool_enrollment.py
"""
from __future__ import annotations

import csv
import re
import unicodedata
import warnings
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SEED_DIR = ROOT / "supabase" / "seed"
ENROLLMENT_WORKBOOKS = [
    ROOT / "Enrollment_Counts.xlsx",
    ROOT / "Enrollment_counts.xlsx",
    ROOT / "data" / "Enrollment_Counts.xlsx",
    ROOT / "data" / "Enrollment_counts.xlsx",
]
SITE_SLUG_ALIASES = {
    "esigcaweni": "sigcaweni",
    "ka-liba": "kaliba",
    "ka-ncesi": "kancesi",
}
FIELDS = ["site_slug", "year", "children_count", "teachers_count", "notes"]


def slugify(text: str) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    slug = text.strip("-")
    return SITE_SLUG_ALIASES.get(slug, slug)


def as_int(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def resolve_workbook() -> Path | None:
    for candidate in ENROLLMENT_WORKBOOKS:
        if candidate.is_file():
            return candidate
    return None


def parse_year_columns(header_row: tuple) -> dict[int, int]:
    """Map calendar year -> column index for Enrollments 'YY headers."""
    year_cols: dict[int, int] = {}
    for idx, cell in enumerate(header_row):
        if not isinstance(cell, str):
            continue
        match = re.search(r"enrollments\s*'(\d{2})", cell.strip(), re.I)
        if match:
            yy = int(match.group(1))
            year = 2000 + yy if yy < 70 else 1900 + yy
            year_cols[year] = idx
    return year_cols


def parse_enrollment_workbook(path: Path) -> list[dict[str, str]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(path, data_only=True)

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []

    year_cols = parse_year_columns(header)
    if not year_cols:
        raise ValueError(f"No Enrollments 'YY columns found in {path}")

    # Town is typically column B (index 1)
    town_idx = 1
    for idx, cell in enumerate(header):
        if isinstance(cell, str) and cell.strip().lower() == "town":
            town_idx = idx
            break

    records: list[dict[str, str]] = []
    for row in rows_iter:
        town = row[town_idx] if len(row) > town_idx else None
        if not isinstance(town, str) or not town.strip():
            continue
        slug = slugify(town)
        if not slug:
            continue

        for year, col_idx in sorted(year_cols.items()):
            count = as_int(row[col_idx] if len(row) > col_idx else None)
            records.append(
                {
                    "site_slug": slug,
                    "year": str(year),
                    "children_count": str(count),
                    "teachers_count": "",
                    "notes": f"Source: {path.name} · {town.strip()}",
                }
            )

    records.sort(key=lambda item: (item["site_slug"], int(item["year"])))
    return records


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    workbook = resolve_workbook()
    if workbook is None:
        raise SystemExit(
            "Enrollment workbook not found. Place Enrollment_Counts.xlsx in the project root."
        )

    rows = parse_enrollment_workbook(workbook)
    out = SEED_DIR / "preschool_enrollment.csv"
    write_csv(out, rows)
    sites = len({row["site_slug"] for row in rows})
    years = len({row["year"] for row in rows})
    print(f"Wrote {len(rows)} enrollment rows ({sites} sites × {years} years) -> {out}")


if __name__ == "__main__":
    main()
