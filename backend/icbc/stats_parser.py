"""
Excel parser for ICBC weekly statistics.

This is a trimmed, backend-friendly version of the logic from church_stats_app.py:
- No Streamlit UI
- No Excel writer
Just parsing each workbook into a list of record dicts.
"""

import io
from typing import Dict, List, Tuple

import openpyxl

KNOWN_CHURCHES = [
    "Bulunga",
    "Engcamini",
    "Gebeni",
    "Gundvwini",
    "Hawane CLC Church",
    "Jubukweni",
    "Kaliba",
    "Kancesi",
    "Lavumisa",
    "Lonhlalane",
    "Lundzi",
    "Lushikishini",
    "Mabhukwini",
    "Makhungutja",
    "Mangcongco",
    "Mantabeni",
    "Mbowane",
    "Mcuba",
    "Mhlabubovu",
    "Mhlangatane",
    "Mlindazwe",
    "Moneni",
    "Mpolonjeni",
    "Mpuluzi",
    "Mshaweni",
    "Ngwempisi",
    "Ngwenya",
    "Njakeni",
    "Nsingweni",
    "Nsubane",
    "Ntondozi",
    "Nyakeni",
    "Nyamane",
    "Nyathela",
    "Phonjwane",
    "Phuzweni",
    "Pine Valley",
    "Sibovini",
    "Sigangeni",
    "Sigcaweni",
    "Sigengeni",
    "Sigombeni",
    "Tibane",
    "Zondwako",
]


def is_church_header(val) -> bool:
    """Return True if this cell value looks like a church name."""
    if not val or not isinstance(val, str):
        return False
    v = val.strip()
    if v.lower().startswith("total icbc"):
        return True
    for name in KNOWN_CHURCHES:
        if name.lower() in v.lower() or v.lower() in name.lower():
            return True
    return False


MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _to_int(value) -> int:
    try:
        return int(value)
    except Exception:
        return 0


def _normalise_date(day_val, mon_val, yr_val) -> str:
    """
    Convert the (day, month, year) values from the sheet into an ISO date string
    YYYY-MM-DD that Django's parse_date can understand.
    """
    day = _to_int(day_val)
    if not day:
        return ""

    year = _to_int(yr_val)
    # Month can be numeric or a month name like "December"
    month_int = 0
    if mon_val is not None:
        text = str(mon_val).strip()
        if text.isdigit():
            month_int = _to_int(text)
        else:
            month_int = MONTH_MAP.get(text.lower(), 0)

    if not year or not month_int:
        # Fallback to original string if we can't normalise
        return f"{yr_val}-{mon_val}-{day_val}"

    return f"{year:04d}-{month_int:02d}-{day:02d}"


def _parse_sheet(ws, filename: str, sheet_name: str) -> List[Dict]:
    all_rows: List[List] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    records: List[Dict] = []
    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        church_name = row[2] if len(row) > 2 else None

        if not church_name or not isinstance(church_name, str):
            i += 1
            continue

        church_name = church_name.strip()

        if church_name.lower().startswith("total icbc"):
            i += 1
            continue

        if not is_church_header(church_name):
            i += 1
            continue

        try:
            year_row = all_rows[i]
            month_row = all_rows[i + 1]
            day_row = all_rows[i + 2]
        except IndexError:
            i += 1
            continue

        week_dates: List[Tuple[int, int, str, int]] = []
        for col_idx in range(5, len(day_row)):
            day_val = day_row[col_idx]
            mon_val = month_row[col_idx] if col_idx < len(month_row) else None
            yr_val = year_row[col_idx] if col_idx < len(year_row) else None
            if isinstance(day_val, str) and day_val.strip().lower() == "total":
                break
            week_dates.append((col_idx, day_val, mon_val, yr_val))

        stat_data: Dict[str, Dict[int, int]] = {}
        j = i + 3
        block_end = min(j + 12, len(all_rows))
        while j < block_end:
            srow = all_rows[j]
            label_c = srow[2] if len(srow) > 2 else None
            label_d = srow[3] if len(srow) > 3 else None
            label = label_c or label_d
            if label and isinstance(label, str):
                label = label.strip()
                stat_data[label] = {}
                for col_idx in range(5, len(srow)):
                    stat_data[label][col_idx] = srow[col_idx]
            if label_c and isinstance(label_c, str) and is_church_header(label_c) and j > i + 3:
                break
            j += 1

        def get_val(label: str, cix: int) -> int:
            return stat_data.get(label, {}).get(cix) or 0

        for (col_idx, day_val, mon_val, yr_val) in week_dates:
            if day_val == 0 or day_val is None:
                continue

            date_str = _normalise_date(day_val, mon_val, yr_val)
            if not date_str:
                continue

            rec = {
                "Church": church_name,
                "Date": date_str,
                "Source File": filename,
                "Sheet": sheet_name,
                "_dedup_key": f"{church_name.lower().strip()}|{date_str.lower().strip()}",
                "Total Attendance": get_val("Church Attendance:", col_idx)
                or get_val("Church Attendance", col_idx),
                "Men": get_val("Men", col_idx),
                "Women": get_val("Women", col_idx),
                "Youth": get_val("Youth", col_idx),
                "Children": get_val("Children", col_idx),
                "Salvations": get_val("Salvations", col_idx),
                "Baptisms": get_val("Baptisms", col_idx),
                "Home Visits": get_val("Home Visits", col_idx),
                "Meals / Food Packs": get_val("Meals / Food Packs", col_idx),
                "Preschool Attendance": get_val("Preschool Attendance", col_idx),
            }
            records.append(rec)

        i = j

    return records


def parse_workbook(file_bytes: bytes, filename: str) -> Tuple[List[Dict], int]:
    """
    Parse ALL sheets in a workbook and deduplicate records by (Church, Date).

    Returns (records, duplicate_count).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    seen_keys: Dict[str, Dict] = {}
    duplicates = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for rec in _parse_sheet(ws, filename, sheet_name):
            key = rec["_dedup_key"]
            if key in seen_keys:
                duplicates += 1
            else:
                seen_keys[key] = rec

    records: List[Dict] = []
    for rec in seen_keys.values():
        rec.pop("_dedup_key", None)
        records.append(rec)

    return records, duplicates

